import cv2
import torch
import numpy as np
import customtkinter as ctk
from PIL import Image
import dlib
from scipy.spatial import distance as dist
import threading
import time
import os
import datetime
import platform
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import face_recognition

# EAR threshold & frame count
EAR_THRESHOLD = 0.25
CONSEC_FRAMES = 20
FRAME_SKIP_FACE_RECOG = 10

frame_counter = 0
drowsy = False
alert_active = False
last_alert_time = 0
drowsiness_events = []
frame_count = 0
last_known_name = "Unknown"

# Setup snapshot & log directories
os.makedirs("snapshots", exist_ok=True)
os.makedirs("known_faces", exist_ok=True)
excel_filename = "session_log.xlsx"

# Excel log setup
if not os.path.exists(excel_filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Drowsiness Log"
    ws.append(["Timestamp", "EAR", "Name"])
    bold_font = Font(bold=True)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=3):
        for cell in col:
            cell.font = bold_font
    wb.save(excel_filename)

# Load known faces
known_face_encodings = []
known_face_names = []
for file in os.listdir("known_faces"):
    if file.endswith(('.jpg', '.png', '.jpeg')):
        img = face_recognition.load_image_file(os.path.join("known_faces", file))
        encodings = face_recognition.face_encodings(img)
        if encodings:
            known_face_encodings.append(encodings[0])
            known_face_names.append(os.path.splitext(file)[0])

# Initialize dlib face landmark predictor
detector = dlib.get_frontal_face_detector()
predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")

LEFT_EYE = list(range(36, 42))
RIGHT_EYE = list(range(42, 48))

def eye_aspect_ratio(eye):
    A = dist.euclidean(eye[1], eye[5])
    B = dist.euclidean(eye[2], eye[4])
    C = dist.euclidean(eye[0], eye[3])
    return (A + B) / (2.0 * C)

# YOLOv5-face setup
from models.experimental import attempt_load
from utils.general import non_max_suppression, scale_coords
from utils.torch_utils import select_device

device = select_device('')
model = attempt_load('weights/yolov5s-face.pt', map_location=device)
model.eval()

# UI setup
ctk.set_appearance_mode("dark")
app = ctk.CTk()
app.title("Drowsiness Detection with Face Recognition")
app.geometry("1000x700")

status_label = ctk.CTkLabel(app, text="Status: ", font=("Arial", 20))
status_label.pack(pady=10)

canvas = ctk.CTkLabel(app, text="")
canvas.pack(pady=10)

ear_display = ctk.CTkLabel(app, text="EAR: 0.00", font=("Arial", 16))
ear_display.pack(pady=4)

name_display = ctk.CTkLabel(app, text="Name: Unknown", font=("Arial", 16))
name_display.pack(pady=4)

# EAR Plot
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

graph_frame = ctk.CTkFrame(app)
graph_frame.pack(pady=10)

ear_values = []

fig, ax = plt.subplots(figsize=(6, 2))
line, = ax.plot([], [], color='lime')
ax.set_ylim(0, 0.4)
ax.set_xlim(0, 50)
ax.set_title("Live EAR")
ax.set_ylabel("EAR")
ax.set_xlabel("Frame")
ax.grid(True, linestyle='--', alpha=0.3)
fig.tight_layout()

canvas_plot = FigureCanvasTkAgg(fig, master=graph_frame)
canvas_plot.get_tk_widget().pack()

def update_plot(ear):
    ear_values.append(ear)
    if len(ear_values) > 50:
        ear_values.pop(0)
    line.set_data(range(len(ear_values)), ear_values)
    ax.set_xlim(max(0, len(ear_values) - 50), len(ear_values))
    ax.set_ylim(0, max(0.3, max(ear_values) + 0.05))
    canvas_plot.draw()

def play_alert():
    global alert_active, last_alert_time
    now = time.time()
    if now - last_alert_time < 5:
        return
    alert_active = True
    last_alert_time = now
    if platform.system() == "Windows":
        import winsound
        winsound.Beep(1000, 1000)
    else:
        os.system('play -nq -t alsa synth 1 sine 1000')  # Linux
    alert_active = False

def save_snapshot(frame, ear, name):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filename = f"snapshots/drowsy_{timestamp.replace(':', '').replace(' ', '_')}.jpg"
    cv2.imwrite(filename, frame)

    # Log to Excel
    wb = load_workbook(excel_filename)
    ws = wb.active
    ws.append([timestamp, round(ear, 3), name])
    wb.save(excel_filename)

# Webcam
cap = cv2.VideoCapture(0)

def process_frame():
    global frame_counter, drowsy, frame_count, last_known_name

    ret, frame = cap.read()
    if not ret:
        app.after(10, process_frame)
        return

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = detector(gray)

    status_text = "AWAKE"
    status_color = (0, 255, 0)

    name = last_known_name

    if frame_count % FRAME_SKIP_FACE_RECOG == 0:
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb)
        face_encodings = face_recognition.face_encodings(rgb, face_locations)

        recognized = False
        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            if any(matches):
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                    recognized = True
                    break

        if recognized:
            last_known_name = name

    name_display.configure(text=f"Name: {last_known_name}")

    for face in faces:
        shape = predictor(gray, face)
        shape_np = np.zeros((68, 2), dtype="int")
        for i in range(68):
            shape_np[i] = (shape.part(i).x, shape.part(i).y)

        left_eye = shape_np[LEFT_EYE]
        right_eye = shape_np[RIGHT_EYE]
        leftEAR = eye_aspect_ratio(left_eye)
        rightEAR = eye_aspect_ratio(right_eye)
        ear = (leftEAR + rightEAR) / 2.0

        update_plot(ear)
        ear_display.configure(text=f"EAR: {ear:.3f}")

        if ear < EAR_THRESHOLD:
            frame_counter += 1
            if frame_counter >= CONSEC_FRAMES and not drowsy:
                drowsy = True
                drowsiness_events.append(datetime.datetime.now())
                threading.Thread(target=play_alert, daemon=True).start()
                save_snapshot(frame, ear, last_known_name)
        else:
            frame_counter = 0
            drowsy = False

        for (x, y) in np.concatenate((left_eye, right_eye), axis=0):
            cv2.circle(frame, (x, y), 2, (255, 255, 0), -1)

    if drowsy:
        status_text = "DROWSY"
        status_color = (0, 0, 255)

    status_label.configure(text=f"Status: {status_text}", text_color=("green" if not drowsy else "red"))
    cv2.putText(frame, f"Status: {status_text}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, status_color, 2)

    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img_pil = Image.fromarray(frame_rgb)
    ctk_img = ctk.CTkImage(light_image=img_pil, size=(640, 480))
    canvas.configure(image=ctk_img)
    canvas.image = ctk_img

    frame_count += 1
    app.after(10, process_frame)

# Start detection loop
app.after(0, process_frame)
app.mainloop()

# Release camera
cap.release()
cv2.destroyAllWindows()

# Final session summary in Excel
wb = load_workbook(excel_filename)
ws = wb.active
ws.append(["Total Drowsiness Events", len(drowsiness_events)])
ws.append([])
wb.save(excel_filename)